import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_gov_briefing_excel():
    # 1. Initialize Workbook and Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Governance Brief"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # 2. Design System Styles (Royal Government Blue theme)
    font_name = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_name, size=15, bold=True, color="FFFFFF")
    section_font = Font(name=font_name, size=12, bold=True, color="1B365D")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    card_title_font = Font(name=font_name, size=9, bold=True, color="595959")
    card_value_font = Font(name=font_name, size=10, bold=True, color="1F4E78")
    data_font = Font(name=font_name, size=10, color="000000")
    data_font_bold = Font(name=font_name, size=10, bold=True, color="000000")
    
    # Fills
    title_fill = PatternFill(start_color="112233", end_color="112233", fill_type="solid") # Very Dark Blue-Black
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Slate Accent Fill
    header_fill = PatternFill(start_color="36648B", end_color="36648B", fill_type="solid") # Royal Steel Blue
    card_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    # Status Fills & Fonts
    status_styles = {
        "Ready / Completed": {
            "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # Soft Green
            "font": Font(name=font_name, size=10, bold=True, color="375623")
        },
        "In Customization": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), # Soft Yellow
            "font": Font(name=font_name, size=10, bold=True, color="833C0C")
        },
        "Operational": {
            "fill": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"), # Soft Blue
            "font": Font(name=font_name, size=10, bold=True, color="1F4E78")
        }
    }
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    thick_bottom_side = Side(border_style="medium", color="1B365D")
    double_bottom_side = Side(border_style="double", color="1B365D")
    
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_section = Border(bottom=double_bottom_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)
    
    # Alignments
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # ==========================================
    # TITLE BLOCK (Rows 1-2)
    # ==========================================
    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = "APCRDA Offline TDR Bond Migration System — Governance & Process Guide"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = align_center_wrap
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    
    # ==========================================
    # GOV BRIEF SUMMARY CARDS (Rows 4-5)
    # ==========================================
    summary_cards = [
        {"range_lbl": "A4:B4", "range_val": "A5:B5", "lbl": "Target Objective", "val": "Migrate Offline TDR Bonds to Secure Ledger"},
        {"range_lbl": "C4:D4", "range_val": "C5:D5", "lbl": "Primary Governance Strategy", "val": "5-Level Verifiable Approval Pipeline"},
        {"range_lbl": "E4:F4", "range_val": "E5:F5", "lbl": "Audit Compliance", "val": "Tamper-Proof Triple Audit Trail"},
        {"range_lbl": "G4:G4", "range_val": "G5:G5", "lbl": "System Mode", "val": "UAT Preparations"}
    ]
    
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 22
    
    for card in summary_cards:
        # Merge label cells
        cells_lbl = card["range_lbl"].split(":")
        if len(cells_lbl) > 1:
            ws.merge_cells(card["range_lbl"])
        lbl_cell = ws[cells_lbl[0]]
        lbl_cell.value = card["lbl"]
        lbl_cell.font = card_title_font
        lbl_cell.fill = card_fill
        lbl_cell.alignment = align_center_wrap
        lbl_cell.border = border_all
        
        # Merge value cells
        cells_val = card["range_val"].split(":")
        if len(cells_val) > 1:
            ws.merge_cells(card["range_val"])
        val_cell = ws[cells_val[0]]
        val_cell.value = card["val"]
        val_cell.font = card_value_font
        val_cell.fill = card_fill
        val_cell.alignment = align_center_wrap
        val_cell.border = border_all

        # Draw borders for merged regions
        cols = range(openpyxl.utils.column_index_from_string(cells_lbl[0][0]), openpyxl.utils.column_index_from_string(cells_lbl[-1][0]) + 1)
        for c in cols:
            ws.cell(row=4, column=c).border = border_all
            ws.cell(row=5, column=c).border = border_all

    # ==========================================
    # SECTION A: THE 5-LEVEL GOVERNANCE PIPELINE
    # ==========================================
    ws.merge_cells("A7:G7")
    sec_hdr = ws["A7"]
    sec_hdr.value = "A. 5-LEVEL OFFICIAL GOVERNANCE & APPROVAL CHAIN"
    sec_hdr.font = section_font
    sec_hdr.fill = section_fill
    sec_hdr.alignment = align_left_wrap
    sec_hdr.border = border_section
    ws.row_dimensions[7].height = 28
    
    headers_a = ["Approval Stage", "Government Authority Role", "Operational Responsibility", "Key Verification Checks", "Authority Bounds & Security", "Digital Signing Method", "System Readiness"]
    ws.row_dimensions[8].height = 24
    for idx, h in enumerate(headers_a, 1):
        cell = ws.cell(row=8, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center_wrap
        cell.border = border_header
        
    data_a = [
        ["Level 0: Entry", "DEO / Surveyor", "Initiates details entry, compiles offline document scans, and uploads land records.", "Matches physical bond copies, cross-checks holder details, and verifies land survey boundaries.", "Can only enter data and modify drafts; cannot approve or release bonds. Limited to their assigned administrative district.", "Secure session login + document hashes", "Ready / Completed"],
        ["Level 1: Validation", "Deputy Tahsildar / Tahsildar", "First level validation. Reviews data validity and verifies survey coordinates against GIS archives.", "Verifies farmer Aadhaar-linked ownership, GIS boundaries, and confirms deed matches offline records.", "Authorized to approve, reject, or return to DEO for edits. Restricted strictly to their mandal/district bounds.", "Aadhaar OTP + System Signature", "Ready / Completed"],
        ["Level 2: Revenue Check", "Special Deputy Collector (SDC)", "Conducts comprehensive revenue audit and verifies historical land acquisition rewards.", "Ensures compliance with Land Pooling Scheme (LPS) criteria, checks for land litigation, and validates ratios.", "Authorized to approve or return. Cannot modify entry details directly to maintain document integrity.", "Aadhaar OTP + System Signature", "Ready / Completed"],
        ["Level 3: Clearance", "Director of Lands", "Performs land allotment clearance checks and verifies plot availability in APCRDA records.", "Validates master plans, returnable plot allocations, and asserts TDR ratio conforms to regulatory guidelines.", "Ensures compliance with urban land laws. Decisions are permanently locked to their administrative ID.", "Aadhaar OTP + System Signature", "Ready / Completed"],
        ["Level 4: Release", "Additional Commissioner / Commissioner", "Final executive approval, signs the digital certificate, and authorizes the bond issuance.", "Reviews the validation timeline, checks validation audits, and mints the immutable TDR certificate.", "Holds the single authority to issue and activate the digital TDR bond. Triggers blockchain certificate minting.", "Aadhaar OTP + Commissioner Digital Signature", "Ready / Completed"]
    ]
    
    current_row = 9
    for row_idx, row_data in enumerate(data_a, current_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            if col_idx == 7:
                cell.alignment = align_center_wrap
                if val in status_styles:
                    cell.fill = status_styles[val]["fill"]
                    cell.font = status_styles[val]["font"]
            elif col_idx in [1, 2]:
                cell.alignment = align_left_wrap
                cell.font = data_font_bold
            else:
                cell.alignment = align_left_wrap
                
            if col_idx != 7 and row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row = row_idx + 1

    # ==========================================
    # SECTION B: 3-PHASE DATA ENTRY & INTEGRATION
    # ==========================================
    current_row += 1
    ws.merge_cells(f"A{current_row}:G{current_row}")
    sec_hdr = ws[f"A{current_row}"]
    sec_hdr.value = "B. 3-PHASE SECURE DATA ENTRY PROCESS"
    sec_hdr.font = section_font
    sec_hdr.fill = section_fill
    sec_hdr.alignment = align_left_wrap
    sec_hdr.border = border_section
    ws.row_dimensions[current_row].height = 28
    
    current_row += 1
    headers_b = ["Form Phase", "Process Objective", "Captured Data Elements", "Automated Validation / Integration", "Internal Security Protection", "User Interface Detail", "System Readiness"]
    ws.row_dimensions[current_row].height = 24
    for idx, h in enumerate(headers_b, 1):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center_wrap
        cell.border = border_header
        
    data_b = [
        ["Phase 1: Holder Profile", "Captures legal identity details of the TDR applicant.", "Name, Son/Daughter/Wife of, door number, street, village, district, email, Aadhaar-registered phone.", "KYC OTP checks verify Aadhaar numbers. Prevents identity mismatches.", "Farmer's Aadhaar number is hashed using HMAC and encrypted in database; never stored in raw readable format.", "Interactive form fields with OTP check trigger.", "In Customization"],
        ["Phase 2: Land & Survey", "Captures geological boundaries and acquisition details of the surrendered land.", "Village name, survey number, ownership deed details, surrendered area (Sq Yds), TDR ratio (e.g. 1:1), plot code.", "Connects to GIS mapping databases to pre-fill survey details and verify exact land extents.", "Area fields are locked to Sq Yards. The system rejects modifications to official GIS pre-filled boundaries.", "GIS validation hooks and pre-fill buttons.", "In Customization"],
        ["Phase 3: File Uploads", "Uploads legal proof documents, scans, and sketches.", "Ownership deeds, Aadhaar card copies, plot allotments, surveyor sketch, and old offline TDR bond copy.", "Calculates a unique SHA-256 hash of each uploaded PDF file to serve as a cryptographic seal.", "Protects uploads against tampering. If any file is modified on the server, the hash check instantly fails.", "File upload zones supporting size checks.", "In Customization"]
    ]
    
    current_row += 1
    for row_idx, row_data in enumerate(data_b, current_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            if col_idx == 7:
                cell.alignment = align_center_wrap
                if val in status_styles:
                    cell.fill = status_styles[val]["fill"]
                    cell.font = status_styles[val]["font"]
            elif col_idx == 1:
                cell.alignment = align_left_wrap
                cell.font = data_font_bold
            else:
                cell.alignment = align_left_wrap
                
            if col_idx != 7 and row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row = row_idx + 1

    # ==========================================
    # SECTION C: SECURITY & TAMPER PREVENTION
    # ==========================================
    current_row += 1
    ws.merge_cells(f"A{current_row}:G{current_row}")
    sec_hdr = ws[f"A{current_row}"]
    sec_hdr.value = "C. ADMINISTRATIVE SECURITY & TAMPER-PREVENTION CONTROLS"
    sec_hdr.font = section_font
    sec_hdr.fill = section_fill
    sec_hdr.alignment = align_left_wrap
    sec_hdr.border = border_section
    ws.row_dimensions[current_row].height = 28
    
    current_row += 1
    headers_c = ["Security Guardrail", "Security Objective", "How it Protects the Department", "Plain-Language Explanation", "Audit Layer", "Administrative Check", "Status"]
    ws.row_dimensions[current_row].height = 24
    for idx, h in enumerate(headers_c, 1):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center_wrap
        cell.border = border_header
        
    data_c = [
        ["Decentralized Ledger", "Prevents database tampering and unauthorized insertions.", "Ensures once a bond is approved or certificate is generated, the record cannot be altered or deleted, even by DB admins.", "Information is stored on a secure, multi-party ledger. Any database changes must match the blockchain record.", "Hyperledger Fabric ledger", "Integrity check compares database states directly with blockchain transactions.", "Operational"],
        ["Tamper-Proof Audit Chain", "Detects internal database manipulation.", "Traces every status change and keeps a historical record. If any entry is modified, deleted, or backdated, the hash chain breaks.", "Records are chained like links. Modifying a single link breaks all subsequent links, alerting administrators.", "PostgreSQL hash-chain", "Health check scripts verify the audit database integrity sequentially.", "Operational"],
        ["Role-Based Access (PDP)", "Prevents unauthorized operations and reviews.", "Ensures surveyors cannot approve bonds, officials cannot edit details, and validators can only process bonds within their district.", "Enforces policy rules based on employee roles. A Surveyor cannot access Tahiti level tasks.", "Cerbos Access PDP Engine", "Enforces access rules on all API calls, logging denials in the audit log.", "Operational"],
        ["Cryptographic Approvals", "Guarantees official signature authenticity.", "Binds approval details (actor, time, decision) using cryptographically signed signatures. Prevents forgery.", "Calculates a signature based on server secret keys. If details are modified, the signature check fails.", "Approval Signatures (HMAC)", "Verify signatures on each stage of the review queue.", "Operational"]
    ]
    
    current_row += 1
    for row_idx, row_data in enumerate(data_c, current_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            if col_idx == 7:
                cell.alignment = align_center_wrap
                if val in status_styles:
                    cell.fill = status_styles[val]["fill"]
                    cell.font = status_styles[val]["font"]
            elif col_idx == 1:
                cell.alignment = align_left_wrap
                cell.font = data_font_bold
            else:
                cell.alignment = align_left_wrap
                
            if col_idx != 7 and row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row = row_idx + 1

    # ==========================================
    # SECTION D: UTILITY & PUBLIC TRUST FEATURES
    # ==========================================
    current_row += 1
    ws.merge_cells(f"A{current_row}:G{current_row}")
    sec_hdr = ws[f"A{current_row}"]
    sec_hdr.value = "D. UTILITY & PUBLIC TRUST FEATURES"
    sec_hdr.font = section_font
    sec_hdr.fill = section_fill
    sec_hdr.alignment = align_left_wrap
    sec_hdr.border = border_section
    ws.row_dimensions[current_row].height = 28
    
    current_row += 1
    headers_d = ["Utility Feature", "Target Beneficiary", "How it Works / User Benefit", "Operational Importance", "Administrative Oversight", "Localization Detail", "Status"]
    ws.row_dimensions[current_row].height = 24
    for idx, h in enumerate(headers_d, 1):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center_wrap
        cell.border = border_header
        
    data_d = [
        ["Public QR Verification", "General Public, Registrar, Banks", "Scanning the certificate QR code redirects users to a public verification page showing authenticity state.", "Allows registrars and banks to immediately verify TDR bonds, eliminating fake or forged paper certificates.", "Admin handles revocations. System updates public pages instantly.", "Telugu / English validation status check.", "Operational"],
        ["Farmer PWA Portal", "Farmers / Land Owners", "Allows farmers to login via Aadhaar OTP on their phone, check validation status, and download certificates.", "Improves transparency. Farmers know exactly which official is currently holding their validation file.", "Admins monitor active sessions and verify KYC phone logins.", "Translates features to Telugu for rural usability.", "Operational"],
        ["Certificate PDF Engine", "Farmers and Officials", "Generates high-definition PDF certificates containing Telugu localization and digital signatures.", "Replaces manual hand-signed papers. Speeds up delivery process.", "Digitally signed using the Commissioner's cryptographic key.", "Includes official AP government emblems.", "Operational"]
    ]
    
    current_row += 1
    for row_idx, row_data in enumerate(data_d, current_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            if col_idx == 7:
                cell.alignment = align_center_wrap
                if val in status_styles:
                    cell.fill = status_styles[val]["fill"]
                    cell.font = status_styles[val]["font"]
            elif col_idx == 1:
                cell.alignment = align_left_wrap
                cell.font = data_font_bold
            else:
                cell.alignment = align_left_wrap
                
            if col_idx != 7 and row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row = row_idx + 1

    # ==========================================
    # AUTO-FIT COLUMN WIDTHS & STYLING POLISH
    # ==========================================
    max_widths = {
        1: 22,  # Column A
        2: 28,  # Column B
        3: 45,  # Column C
        4: 55,  # Column D
        5: 35,  # Column E
        6: 25,  # Column F
        7: 20   # Column G
    }
    
    for col_idx in range(1, 8):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for r in range(7, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        
        calc_width = max(max_len + 3, 12)
        target_width = min(calc_width, max_widths.get(col_idx, 50))
        ws.column_dimensions[col_letter].width = target_width
        
    output_filename = "APCRDA_TDR_Government_Briefing.xlsx"
    wb.save(output_filename)
    print(f"Government Briefing Excel generated successfully: {output_filename}")

if __name__ == "__main__":
    create_gov_briefing_excel()
