import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_status_excel():
    # 1. Initialize Workbook and Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TDR System Flow & Status"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # 2. Design System Styles
    font_name = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_name, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_name, size=13, bold=True, color="1F4E78")
    header_font = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    card_title_font = Font(name=font_name, size=10, bold=True, color="595959")
    card_value_font = Font(name=font_name, size=10, bold=True, color="1F4E78")
    data_font = Font(name=font_name, size=10, color="000000")
    data_font_bold = Font(name=font_name, size=10, bold=True, color="000000")
    
    # Fills
    title_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Dark Navy
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Light Slate Blue
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Steel Blue
    card_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Very Light Gray
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Off-white zebra
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Status Fills & Fonts
    status_styles = {
        "Completed": {
            "fill": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"), # Light Green
            "font": Font(name=font_name, size=10, bold=True, color="375623")
        },
        "In Progress": {
            "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"), # Light Yellow/Orange
            "font": Font(name=font_name, size=10, bold=True, color="833C0C")
        },
        "Pending": {
            "fill": PatternFill(start_color="E6EDF5", end_color="E6EDF5", fill_type="solid"), # Light Blue-Gray
            "font": Font(name=font_name, size=10, bold=True, color="1F4E78")
        }
    }
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thick_bottom_side = Side(border_style="medium", color="1B365D")
    double_bottom_side = Side(border_style="double", color="1B365D")
    
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_section = Border(bottom=double_bottom_side)
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    
    # Alignments
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right_wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    # ==========================================
    # TITLE BLOCK (Rows 1-2)
    # ==========================================
    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = "APCRDA Offline TDR Bond Migration — System Development Status & Workflow Guide"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = align_center_wrap
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    
    # ==========================================
    # PROJECT SUMMARY CARDS (Rows 4-5)
    # ==========================================
    # Column mappings: A & B, C & D, E & F, G & H
    summary_cards = [
        {"range_label": "A4:B4", "range_val": "A5:B5", "label": "Project Name", "val": "APCRDA TDR Bond Migration"},
        {"range_label": "C4:D4", "range_val": "C5:D5", "label": "Tech Stack Focus", "val": "Next.js 14, Cerbos, Fabric, Prisma"},
        {"range_label": "E4:F4", "range_val": "E5:F5", "label": "Active Core Files", "val": "128 Files"},
        {"range_label": "G4:G4", "range_val": "G5:G5", "label": "Project Status", "val": "Week 3 (Approval Phase)"}
    ]
    
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 22
    
    for card in summary_cards:
        # Merge label cells
        cells_lbl = card["range_label"].split(":")
        if len(cells_lbl) > 1:
            ws.merge_cells(card["range_label"])
        lbl_cell = ws[cells_lbl[0]]
        lbl_cell.value = card["label"]
        lbl_cell.font = card_title_font
        lbl_cell.fill = card_fill
        lbl_cell.alignment = align_center_wrap
        lbl_cell.border = border_all
        
        # Merge value cells
        cells_val = card["range_val"].split(":")
        if len(cells_val) > 1:
            ws.merge_cells(card["range_val"])
        val_cell = ws[cells_val[0]]
        val_cell.value = card["val"]
        val_cell.font = card_value_font
        val_cell.fill = card_fill
        val_cell.alignment = align_center_wrap
        val_cell.border = border_all

        # Apply borders to all merged cells
        cols = range(openpyxl.utils.column_index_from_string(cells_lbl[0][0]), openpyxl.utils.column_index_from_string(cells_lbl[-1][0]) + 1)
        for c in cols:
            ws.cell(row=4, column=c).border = border_all
            ws.cell(row=5, column=c).border = border_all

    # ==========================================
    # SECTION 1: ARCHITECTURE & TECH STACK
    # ==========================================
    # Header
    ws.merge_cells("A7:G7")
    sec1_hdr = ws["A7"]
    sec1_hdr.value = "1. SYSTEM ARCHITECTURE & TECHNOLOGY STACK"
    sec1_hdr.font = section_font
    sec1_hdr.fill = section_fill
    sec1_hdr.alignment = align_left_wrap
    sec1_hdr.border = border_section
    ws.row_dimensions[7].height = 28
    
    # Table Column Headers
    tech_headers = ["Layer", "Technology / Framework", "Version / Spec", "Core Responsibility & Scope", "Security & Integration Handshake", "Status", "Implementation Status Notes"]
    ws.row_dimensions[8].height = 24
    for idx, h in enumerate(tech_headers, 1):
        cell = ws.cell(row=8, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center_wrap
        cell.border = border_header
        
    # Table Data
    tech_data = [
        ["Core Web Application", "Next.js 14 (App Router)", "React 18 LTS", "Server-rendered user interfaces, routing structure, React-Hook-Form bindings, pages for all portals.", "Supabase Session Cookies validation inside Next.js Middleware with automatic renewals.", "Completed", "Framework and layout shells are fully operational; global middleware redirects inactive users."],
        ["Identity & Authentication", "Supabase Auth", "v2.108.2 client", "Authenticates farmers via OTP and government officials via employee credentials & simulated NIC SSO.", "Generates cryptographically signed JWT tokens parsed on the API routes viagetCurrentUser helper.", "Completed", "APIs for farmer OTP requests, official credentials matching, and login page wrappers are completed."],
        ["Granular Authorization", "Cerbos PDP", "v0.31.0 gRPC client", "Policy-as-code middleware implementing ALLOW/DENY check on each backend route based on user roles and rules.", "Queries Cerbos PDP locally via gRPC, evaluates policies (e.g. resource_bond.yaml), and records a unique cerbosCallId.", "Completed", "5 major YAML policies (bond, approval, certificate, document, derived roles) compiled and running in Docker."],
        ["Database & ORM", "Prisma & PostgreSQL", "Prisma 5.22, PG 15", "Object-Relational Mapping (ORM) and persistent data storage. Implements tables for bonds, approvals, and logs.", "Next.js API routes query the PG database using Prisma Client. Database enforces row-level security constraints.", "Completed", "Database models seeded successfully with test records (farmers, officials, villages). Schema push active."],
        ["Immutable Audit Trail", "Hyperledger Fabric", "v2.5 Gateway SDK", "Decentralized blockchain ledger storing immutable records of bond creation, approvals, and certificate hashes.", "Next.js API routes connect via Fabric Gateway, execute chaincode write transactions and record the fabricTxId.", "Completed", "Chaincode deployment configured to run locally. SDK client handles transactions, fallback mock mode available."],
        ["Cryptographic Integrity", "Custom HMAC Security", "SHA-256 Engine", "Generates cryptographic signature hashes for approval steps to construct the audit trail's tamper-proof chain.", "Calculates signature = HMAC(prevHash + bondId + actorId + decision) using secure server-side environment keys.", "Completed", "HMAC module implemented and validated using unit tests. Used to verify that approval history is authentic."],
        ["Document Generation", "PDFKit", "v0.19.1 Engine", "Generates high-definition PDF certificates containing Telugu internationalization and official signature marks.", "Mints certificate details, generates a secure verification QR code, and outputs a downloadable PDF stream.", "Completed", "PDF layout template defined with QR check. Successfully integrates Commissioner validation hashes."],
        ["Data Entry Frontend", "React / Next.js Forms", "Next.js Pages", "Multi-phase data collection forms (holder details, land coordinates, document uploads) for DEO surveyors.", "Validates inputs on client and server using Zod schemas, fetches prefill values, and handles document uploads.", "In Progress", "Currently refining form validation errors and UI responsiveness in Phase 1, Phase 2, and Phase 3 forms."]
    ]
    
    current_row = 9
    for row_idx, row_data in enumerate(tech_data, current_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            # Formatting for Status column
            if col_idx == 6:
                status = val
                cell.alignment = align_center_wrap
                if status in status_styles:
                    cell.fill = status_styles[status]["fill"]
                    cell.font = status_styles[status]["font"]
            elif col_idx in [1, 2, 3]:
                cell.alignment = align_left_wrap
                cell.font = data_font_bold
            else:
                cell.alignment = align_left_wrap
                
            # Zebra striping (except status column)
            if col_idx != 6 and row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row = row_idx + 1
        
    # ==========================================
    # SECTION 2: CODEBASE MODULES
    # ==========================================
    current_row += 1
    ws.merge_cells(f"A{current_row}:G{current_row}")
    sec2_hdr = ws[f"A{current_row}"]
    sec2_hdr.value = "2. CODEBASE FILE TREE & IMPLEMENTATION STATUS"
    sec2_hdr.font = section_font
    sec2_hdr.fill = section_fill
    sec2_hdr.alignment = align_left_wrap
    sec2_hdr.border = border_section
    ws.row_dimensions[current_row].height = 28
    
    current_row += 1
    code_headers = ["Module Area", "Relative File Path", "Module Type", "Core Functionality / Responsibility", "Active Features & Logic", "Status", "Developer / Development Notes"]
    ws.row_dimensions[current_row].height = 24
    for idx, h in enumerate(code_headers, 1):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center_wrap
        cell.border = border_header
        
    code_data = [
        # DB & Auth Configuration
        ["Database Schema", "prisma/schema.prisma", "ORM Schema", "Declares all relational models: TdrBond, BondHolder, BondLandDetail, BondDocument, ApprovalStep, Official, Farmer, AuditLog.", "Defines tables, enums (OfficialRole, BondStatus), foreign keys, and indexes for performant queries.", "Completed", "Fully synchronized with the PostgreSQL database. Used by seed.ts to inject test users."],
        ["Middleware", "src/middleware.ts", "Next.js Edge", "Filters unauthorized navigation, inspects JWT validity, handles session refresh, redirects to appropriate portals.", "Implements path matching for /deo/*, /official/*, and /farmer/* routes; redirects unauthenticated visitors to login.", "Completed", "Successfully routing users based on role claims. Logs session updates in the server console."],
        # Cerbos Policies
        ["Authz Rules", "cerbos/policies/derived_roles.yaml", "Cerbos Policy", "Calculates dynamic roles such as bond_owner, l1_approver, l2_approver based on district matching and role level.", "Enables context-sensitive evaluation (e.g., surveyor can edit only if bond status is DRAFT and created by them).", "Completed", "Derived roles policy compiled and validated in Cerbos unit test suites."],
        ["Authz Rules", "cerbos/policies/resource_bond.yaml", "Cerbos Policy", "Restricts actions on the TdrBond resource: view, edit, submit, reject, return, delete based on user role and step state.", "Ensures only authorized actors can progress a bond status (e.g., DEO can delete only DRAFT status).", "Completed", "Full access matrix tested against multiple user roles and boundary conditions."],
        ["Authz Rules", "cerbos/policies/resource_approval.yaml", "Cerbos Policy", "Declares validation policies governing who can perform actions on the ApprovalStep resource.", "Maps levels (1 to 4) directly to roles (Tahsildar, SDC, Director, Commissioner) preventing out-of-order processing.", "Completed", "Compiled successfully, integrated with gRPC enforcement middleware."],
        # Security & Blockchain Modules
        ["Security Module", "src/lib/security/hmac.ts", "Utility", "Computes HMAC hashes using SHA-256 for approval signatures and anonymizes farmer Aadhaar numbers.", "Uses server-only environment secret keys to generate signatures that link sequential approval steps.", "Completed", "Verified by Jest unit tests. Cryptographically guarantees approval history cannot be modified."],
        ["Audit Module", "src/lib/audit.ts", "Logic Helper", "Appends records to the database audit_log table. Chains sequential logs by hashing the previous chain_hash.", "Integrates both Cerbos PDP call ID and Fabric blockchain transaction ID into a unified audit timeline.", "Completed", "Active on state-modifying actions. Standardized formatting enforces trace joins."],
        ["Fabric Client", "src/lib/fabric/client.ts", "Integration", "Initializes the gRPC connection to the Hyperledger Fabric gateway and fetches the smart contract instance.", "Loads crypto credentials (private keys, certificates) from file system paths; falls back to mock logger in dev.", "Completed", "Ready for local or staging Fabric peers. TLS configuration verified for client auth."],
        ["Fabric Gateway", "src/lib/fabric/gateway.ts", "Integration", "Provides high-level TS interfaces: createBond(), recordApproval(), mintCertificate(), queryBondState() to talk to Fabric ledger.", "Submits transactions asynchronously using the Gateway SDK, monitors block commits, and returns transaction hashes.", "Completed", "Features mock fallback logs for teams working without a running local Fabric instance."],
        # Form UI Components (In Progress)
        ["Form Containers", "src/components/bond-form/BondEntryForm.tsx", "Frontend UI", "Parent controller for the 3-phase bond entry wizard. Manages state, handles drafts, and processes submissions.", "Controls step transitions, loads existing draft data, displays step navigation bar, and handles form submission errors.", "In Progress", "Currently adjusting draft auto-save debounce triggers and submission transition loaders."],
        ["Form Components", "src/components/bond-form/Phase1HolderForm.tsx", "Frontend UI", "Phase 1: Captures holder details (name, relation, door number, street, village, Aadhaar phone) and triggers KYC OTP.", "Runs client-side Zod validation, hooks into Aadhaar validation services, and displays OTP verify modal.", "In Progress", "Refining OTP request cooldown timer and Aadhaar format masks for the DEO input fields."],
        ["Form Components", "src/components/bond-form/Phase2LandForm.tsx", "Frontend UI", "Phase 2: Gathers land details (village, survey number, surrendered area, issued extent, TDR ratio, plot code).", "Prefills GIS survey data from the mock GIS server, checks that units are in Sq Yards, and handles ratio fields.", "In Progress", "Improving prefills when surveyor enters survey number, resolving minor state sync glitches."],
        ["Form Components", "src/components/bond-form/DocumentUploadPhase.tsx", "Frontend UI", "Phase 3: Provides PDF file drop zones for required uploads (ownership deeds, sketch, plot allotment, Aadhaar copy).", "Interacts with upload API, tracks upload progress, displays file size limits, and shows preview thumbnails.", "In Progress", "Fixing file drag-and-drop boundary layout issue in Safari and older browser viewports."],
        # Approval Frontend UI
        ["Approval UI", "src/components/approval/BondReviewPanel.tsx", "Frontend UI", "Main workspace for official validators. Displays comprehensive bond information, maps, uploaded files, and action actions.", "Loads approval steps history, renders Reject/Return/Approve actions, triggers OTP signing, and shows validation alerts.", "In Progress", "Adding multi-page PDF document viewer widget within the right side of the review screen."],
        # Farmer Dashboard Component
        ["Farmer Dashboard", "src/components/farmer/BondStatusTracker.tsx", "Frontend UI", "Interactive timeline displaying the current status of the farmer's bond and the steps remaining in the approval chain.", "Listens to realtime updates, highlights completed stages (green) and pending review stages (blue/grey).", "In Progress", "Refining step label translations in Telugu (i18n integration sync)."],
        # API Routes
        ["API Routes", "src/app/api/auth/otp/request/route.ts", "API Route", "Generates a 6-digit OTP, hashes it, stores it in the database, and mocks SMS transmission.", "Validates input phone numbers via Zod, sets 5-minute expiry, and blocks rapid requests.", "Completed", "Operational. Integrated with farmer login and holder KYC forms."],
        ["API Routes", "src/app/api/auth/otp/verify/route.ts", "API Route", "Validates the submitted OTP. On success, issues a Supabase JWT token and sets cookies.", "Matches hash of input with stored OtpRequest, marks OTP as used, and returns profile info.", "Completed", "Tested, handles edge cases (expired OTPs, double-submit attempts, brute-force limits)."],
        ["API Routes", "src/app/api/bonds/[id]/submit/route.ts", "API Route", "Transitions a draft bond status to PENDING_L1 and executes the Fabric ledger entry.", "Requires Cerbos authorization, triggers CreateBond on blockchain, writes to DB and audit log.", "Completed", "Integrated with blockchain network. Supports idempotency key safety headers."],
        ["API Routes", "src/app/api/approvals/[bondId]/approve/route.ts", "API Route", "Records official approval at Level 1, 2, 3, or 4. Generates signature hash and fabric logs.", "Checks Cerbos, verifies OTP signature, calculates HMAC, updates DB status, calls Fabric, updates audit log.", "Completed", "Verified through manual tests. Triggers state changes and signs transactions successfully."],
        ["API Routes", "src/app/api/certificates/[bondId]/generate/route.ts", "API Route", "Creates the final PDF certificate using PDFKit, signs it, uploads to Supabase, and logs on Fabric.", "Fetches bond details, renders PDF layout, stamps Commissioner hash, sends to Fabric ledger.", "Completed", "Successfully outputs standard PDFs with working QR validation links."]
    ]
    
    current_row += 1
    start_data_row = current_row
    for row_idx, row_data in enumerate(code_data, start_data_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            # Formatting for Status column
            if col_idx == 6:
                status = val
                cell.alignment = align_center_wrap
                if status in status_styles:
                    cell.fill = status_styles[status]["fill"]
                    cell.font = status_styles[status]["font"]
            elif col_idx == 1:
                cell.alignment = align_left_wrap
                cell.font = data_font_bold
            elif col_idx == 2:
                cell.alignment = align_left_wrap
                # Clean monospace feel for paths
                cell.font = Font(name="Consolas", size=9, color="333333")
            else:
                cell.alignment = align_left_wrap
                
            # Zebra striping (except status column)
            if col_idx != 6 and row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row = row_idx + 1

    # ==========================================
    # SECTION 3: SYSTEM WORKFLOWS
    # ==========================================
    current_row += 1
    ws.merge_cells(f"A{current_row}:G{current_row}")
    sec3_hdr = ws[f"A{current_row}"]
    sec3_hdr.value = "3. DETAILED SYSTEM WORKFLOWS & DATA FLOWS (ALL IN ONE INTEGRATED FLOW)"
    sec3_hdr.font = section_font
    sec3_hdr.fill = section_fill
    sec3_hdr.alignment = align_left_wrap
    sec3_hdr.border = border_section
    ws.row_dimensions[current_row].height = 28
    
    current_row += 1
    flow_headers = ["Workflow Group", "Step #", "Process Step & Actor", "Action / Process Description", "Files & Endpoints Involved", "Technical Handshake & Mechanisms", "Status"]
    ws.row_dimensions[current_row].height = 24
    for idx, h in enumerate(flow_headers, 1):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center_wrap
        cell.border = border_header
        
    flow_data = [
        # Flow 1: Authentication
        ["1. AUTHENTICATION & SESSION", "1.1", "Request Login OTP (Farmer)", "Farmer submits their phone number. System matches phone against seeded Aadhaar records to verify registration.", "api/auth/otp/request/route.ts", "Zod parsing -> DB check -> Generate OTP -> Hash OTP in DB -> Mock SMS dispatch.", "Completed"],
        ["1.1 SESSION MANAGEMENT", "1.2", "Verify Login OTP & Issue JWT", "Farmer inputs the received 6-digit OTP code. Server verifies, establishes session, and returns cookie.", "api/auth/otp/verify/route.ts", "DB lookup -> Match OTP hashes -> Generate Supabase JWT -> Set httpOnly cookie session.", "Completed"],
        ["1.1 SESSION MANAGEMENT", "1.3", "Official Credentials / SSO Login", "Officials login via simulated National Informatics Centre (NIC) SSO or employee credentials.", "api/auth/official/login/route.ts", "Verify employee credentials against Official table -> Set JWT with claims (role, districtCode).", "Completed"],
        ["1.1 SESSION MANAGEMENT", "1.4", "Middleware Access Control Check", "Next.js middleware interceptor parses the session cookie, checks expiration, and validates routing scopes.", "src/middleware.ts", "Decrypt JWT token -> Extract role claims -> Match routing pattern (e.g. /deo/*) -> Slide session cookie.", "Completed"],
        
        # Flow 2: DEO Data Entry
        ["2. DEO 3-PHASE DATA ENTRY", "2.1", "Create Draft Bond (Phase 1)", "DEO Surveyor inputs basic holder details. Fills name, relationship, door number, and Aadhaar phone.", "deo/bonds/new/page.tsx, Phase1HolderForm.tsx, api/bonds/route.ts", "Checks create permissions via Cerbos PDP -> Hash Aadhaar via HMAC -> Encrypt Aadhaar -> Insert TdrBond (DRAFT).", "In Progress"],
        ["2.2 DATA COLLECTION WIZARD", "2.2", "GIS Prefill & Survey Info (Phase 2)", "DEO enters the survey number and village name. The application pre-fills geographical survey area details.", "Phase2LandForm.tsx, api/bonds/prefill/[surveyNo]/route.ts", "Fetch GIS coordinates from mock external GIS adaptor -> Validate unit is in Sq Yards -> Update TdrBond record.", "In Progress"],
        ["2.2 DATA COLLECTION WIZARD", "2.3", "Document Upload & Hashing (Phase 3)", "DEO uploads ownership deeds, surveyor sketch, and Aadhaar copy as PDF files.", "DocumentUploadPhase.tsx, api/documents/upload/route.ts", "Parse file stream -> Compute SHA-256 hash -> Store PDF in Supabase Storage -> Mock IPFS CID -> Save BondDocument.", "In Progress"],
        ["2.2 DATA COLLECTION WIZARD", "2.4", "Submit Draft Bond for Review", "DEO clicks Submit. The bond undergoes strict final validation and enters the approval pipeline.", "BondEntryForm.tsx, api/bonds/[id]/submit/route.ts", "Zod schema check -> Enforce create permissions -> Write status PENDING_L1 -> Submit Fabric CreateBond -> Chain audit log.", "In Progress"],
        
        # Flow 3: Approval Chain
        ["3. 5-LEVEL APPROVAL CHAIN", "3.1", "Fetch Review Queue (Official)", "Official logs in and views their dashboard queue. The list shows only bonds pending validation at their specific level.", "official/queue/page.tsx, api/approvals/queue/route.ts", "Extract official's role & district from JWT -> Query Prisma for pending items matching role constraints (Level 1-4).", "Completed"],
        ["3.1 APPROVAL PIPELINE", "3.2", "Open Review Workspace", "Official opens the review panel for a specific bond, showing data, GIS mapping preview, and documents.", "official/bonds/[id]/review/page.tsx, BondReviewPanel.tsx", "Query full TdrBond relational details -> Fetch documents from storage -> Render side-by-side workspace.", "In Progress"],
        ["3.1 APPROVAL PIPELINE", "3.3", "Request Approval OTP (Signing)", "Official triggers OTP to their official registered phone number to validate their signature identity.", "api/auth/approval-otp/request/route.ts", "Validate actor identity -> Generate OTP -> Hash OTP in DB -> Mock official SMS transmission.", "Completed"],
        ["3.1 APPROVAL PIPELINE", "3.4", "Process Approval / Transition", "Official enters OTP and submits decision. System validates, signs transaction, and routes to next level.", "api/approvals/[bondId]/approve/route.ts", "Verify OTP -> Call Cerbos PDP -> Compute approval signature HMAC -> Log to Fabric -> Write Postgres audit log.", "Completed"],
        ["3.1 APPROVAL PIPELINE", "3.5", "Return Bond to Previous Step", "If data is incorrect, validator returns the bond to the previous level or back to the DEO for corrections.", "api/approvals/[bondId]/return/route.ts", "Call Cerbos PDP -> Record decision (RETURNED) -> Set status to previous level or DRAFT -> Fabric & Audit log updates.", "Completed"],
        ["3.1 APPROVAL PIPELINE", "3.6", "Reject Bond Permanently", "Official rejects the bond due to mismatching legal records. This terminates the workflow path.", "api/approvals/[bondId]/reject/route.ts", "Call Cerbos PDP -> Record decision (REJECTED) -> Set status to REJECTED -> Fabric ledger record -> Audit log update.", "Completed"],
        
        # Flow 4: Certificate Generation
        ["4. CERTIFICATE GENERATION", "4.1", "Final Level 4 Commissioner Approval", "Commissioner signs and issues final approval, prompting state transition to ACTIVE.", "api/approvals/[bondId]/approve/route.ts", "Transition bond state to ACTIVE -> Trigger certificate generator trigger asynchronously.", "Completed"],
        ["4.1 CERTIFICATE GENERATION", "4.2", "Generate PDF Certificate", "System renders high-quality certificate PDF with Andhra Pradesh emblem and digital QR code.", "lib/pdf/certificate.ts, api/certificates/[bondId]/generate/route.ts", "PDFKit renders document -> Embed layout details -> Generate QR linking to verify/[tdrNumber] -> Stream PDF.", "Completed"],
        ["4.1 CERTIFICATE GENERATION", "4.3", "Sign Certificate & Store File", "System signs PDF with simulated Commissioner digital signature, uploads to storage, and generates hash.", "api/certificates/[bondId]/generate/route.ts", "Calculate SHA-256 of PDF -> Generate signature HMAC -> Write to Supabase bucket -> Fetch IPFS CID mock.", "Completed"],
        ["4.1 CERTIFICATE GENERATION", "4.4", "Mint Certificate on Blockchain", "System binds the generated certificate IPFS CID and signature hash to the blockchain ledger.", "lib/fabric/gateway.ts (mintCertificate)", "Execute MintCertificate transaction on Fabric ledger -> Save fabricTxId and certificateIpfsCid to PostgreSQL.", "Completed"],
        ["4.1 CERTIFICATE GENERATION", "4.5", "Farmer Download / PWA Download", "Farmer logs into PWA, views their issued certificates, and downloads PDF via phone OTP.", "farmer/certificates/[id]/page.tsx, api/certificates/[bondId]/download/route.ts", "Validate OTP -> Fetch PDF from Supabase Storage -> Stream file download to user browser.", "Completed"],
        ["4.1 CERTIFICATE GENERATION", "4.6", "Public QR Code Verification", "Third-parties scan the certificate QR code, opening a public page that checks the ledger's authenticity.", "verify/[tdrNumber]/page.tsx, api/certificates/[bondId]/verify/route.ts", "Read tdrNumber -> Query DB details -> Query Fabric ledger -> Assert hashes match -> Display verified badge.", "Completed"],
        
        # Flow 5: Three-Layer Auditing
        ["5. THREE-LAYER AUDIT INTEGRITY", "5.1", "Record Cerbos Decision Logs", "Every authorization enforce check writes allow/deny decisions with exact policy logic inputs.", "src/lib/cerbos/enforce.ts", "Cerbos PDP engine prints structured decision logs with cerbosCallId to audit file system.", "Completed"],
        ["5.2 AUDITING MECHANISMS", "5.2", "Build Database Hash-Chain", "Every state modification writes a record to the PostgreSQL audit_log table, chaining to the previous log.", "src/lib/audit.ts", "Fetch latest audit log chain_hash -> Compute new chain_hash = SHA256(prevChainHash + data) -> Insert record.", "Completed"],
        ["5.2 AUDITING MECHANISMS", "5.3", "Submit Ledger Record", "Every state transition, approval step signature, and certificate mint is recorded on the Fabric ledger.", "src/lib/fabric/gateway.ts", "Submit transactions to blockchain network -> Store peer-verified fabricTxId inside Database tables.", "Completed"],
        ["5.2 AUDITING MECHANISMS", "5.4", "Run Integrity Scan Check", "Operational scripts scan the PostgreSQL audit logs and cross-reference hashes with the Fabric ledger.", "scripts/pre-launch-check.sh, api/health/route.ts", "Walk DB audit table -> Verify hashes sequence -> Assert DB data aligns with blockchain transactions -> Alert on mismatch.", "Completed"]
    ]
    
    current_row += 1
    start_data_row = current_row
    for row_idx, row_data in enumerate(flow_data, start_data_row):
        ws.row_dimensions[row_idx].height = 24
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = data_font
            cell.border = border_all
            
            # Formatting for Status column
            if col_idx == 7:
                status = val
                cell.alignment = align_center_wrap
                if status in status_styles:
                    cell.fill = status_styles[status]["fill"]
                    cell.font = status_styles[status]["font"]
            elif col_idx == 1:
                cell.alignment = align_left_wrap
                cell.font = data_font_bold
            elif col_idx == 2:
                cell.alignment = align_center_wrap
                cell.font = data_font_bold
            else:
                cell.alignment = align_left_wrap
                
            # Zebra striping (except status column)
            if col_idx != 7 and row_idx % 2 == 1:
                cell.fill = zebra_fill
        current_row = row_idx + 1

    # ==========================================
    # 3. AUTO-FIT COLUMN WIDTHS & STYLING POLISH
    # ==========================================
    max_widths = {
        1: 22,  # Column A: Layer / Module / Flow Group
        2: 32,  # Column B: Tech / File Path / Step #
        3: 28,  # Column C: Spec / File Type / Process Step
        4: 55,  # Column D: Responsibility / Functionality / Description
        5: 50,  # Column E: Handshake / Logic / Endpoints
        6: 15,  # Column F: Status
        7: 55   # Column G: Status Notes / Developer Notes / Technical Mechanism
    }
    
    for col_idx in range(1, 8):
        col_letter = get_column_letter(col_idx)
        # Apply standard auto-fit width but cap at max_widths
        max_len = 0
        # Iterate only through rows 8 and onwards to avoid the merged title cell distorting sizes
        for r in range(7, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        
        # Use length with padding or use default
        calc_width = max(max_len + 3, 12)
        target_width = min(calc_width, max_widths.get(col_idx, 50))
        ws.column_dimensions[col_letter].width = target_width
        
    # Save Workbook
    output_filename = "APCRDA_TDR_System_Flow_and_Status.xlsx"
    wb.save(output_filename)
    print(f"Excel Sheet generated successfully: {output_filename}")

if __name__ == "__main__":
    create_status_excel()
